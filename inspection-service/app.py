from __future__ import annotations

import base64
import copy
import csv
import hashlib
import json
import re
import tempfile
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import cv2
import numpy as np
import streamlit as st
from PIL import Image, ImageDraw, ImageOps
from ultralytics import YOLO


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
ORIGINAL_DIR = DATA_DIR / "originals"
ANNOTATED_DIR = DATA_DIR / "annotated"
LABELS_CSV = DATA_DIR / "review_labels.csv"
EXCEL_REPORT_DIR = DATA_DIR / "excel_reports"
TEMPLATE_DIR = DATA_DIR / "templates"
DESKTOP_DIR = Path.home() / "Desktop"
# 优先使用仓库自带模板，其次才找桌面（跨机器部署时桌面通常没有）
BUNDLED_TEMPLATE_PATH = APP_DIR / "templates" / "万达二店巡检表6.23日.xlsx"
DESKTOP_TEMPLATE_PATH = DESKTOP_DIR / "万达二店巡检表6.23日.xlsx"
INSPECTION_TEMPLATE_PATH = BUNDLED_TEMPLATE_PATH if BUNDLED_TEMPLATE_PATH.exists() else DESKTOP_TEMPLATE_PATH
DEFAULT_STORE_NAME = "万达二店"
PHOTO_GRID_COLUMNS = ["A", "D", "G", "J"]
PHOTO_GRID_START_ROW = 16
PHOTO_GRID_ROW_STEP = 22
PHOTO_DISPLAY_SIZE = (240, 320)
PHOTO_CANVAS_SIZE = (900, 1200)
TRAINED_MODEL_PATH = APP_DIR / "models" / "floor_litter_seed_best.pt"
DEFAULT_MODEL_PATH = str(TRAINED_MODEL_PATH) if TRAINED_MODEL_PATH.exists() else "yolo11n.pt"
HAS_TRAINED_MODEL = TRAINED_MODEL_PATH.exists()
DEFAULT_CONFIDENCE = 0.12 if HAS_TRAINED_MODEL else 0.30
DEFAULT_IMAGE_SIZE = 960 if HAS_TRAINED_MODEL else 640
DEFAULT_FLOOR_START_PERCENT = 10 if HAS_TRAINED_MODEL else 40
TRAINED_FLOOR_CLASSES = {"paper_scrap", "stain"}

DEFAULT_TARGET_CLASSES = (
    "paper_scrap, stain, corner_dust, dust, floor_litter, light_floor_litter, colored_floor_litter, "
    "dirty, trash, garbage, waste, oil, grease, residue, mold, "
    "food residue, bottle, cup, bowl, fork, knife, spoon, banana, apple, "
    "orange, sandwich, hot dog, pizza, donut, cake"
)

LITTER_SENSITIVITY = {
    "低": {
        "sat_threshold": 45,
        "bright_delta": 26,
        "min_area": 80,
        "max_area": 900,
        "max_side": 80,
        "max_cluster_side": 85,
        "colored_min_value": 95,
    },
    "中": {
        "sat_threshold": 35,
        "bright_delta": 22,
        "min_area": 45,
        "max_area": 1000,
        "max_side": 75,
        "max_cluster_side": 95,
        "colored_min_value": 80,
    },
    "高": {
        "sat_threshold": 28,
        "bright_delta": 18,
        "min_area": 20,
        "max_area": 1200,
        "max_side": 95,
        "max_cluster_side": 120,
        "colored_min_value": 65,
    },
}

DUST_SENSITIVITY = {
    "低": {
        "delta": 16,
        "max_gray": 205,
        "very_dark": 62,
        "min_window_pixels": 1400,
        "min_window_density": 0.024,
        "score_percentile": 92,
        "max_region_ratio": 0.34,
    },
    "中": {
        "delta": 12,
        "max_gray": 215,
        "very_dark": 70,
        "min_window_pixels": 900,
        "min_window_density": 0.018,
        "score_percentile": 88,
        "max_region_ratio": 0.42,
    },
    "高": {
        "delta": 9,
        "max_gray": 225,
        "very_dark": 82,
        "min_window_pixels": 600,
        "min_window_density": 0.012,
        "score_percentile": 84,
        "max_region_ratio": 0.50,
    },
}


def ensure_dirs() -> None:
    ORIGINAL_DIR.mkdir(parents=True, exist_ok=True)
    ANNOTATED_DIR.mkdir(parents=True, exist_ok=True)
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


@st.cache_resource(show_spinner=False)
def load_model(model_path: str) -> YOLO:
    return YOLO(model_path)


def default_template_path() -> Path | None:
    return INSPECTION_TEMPLATE_PATH if INSPECTION_TEMPLATE_PATH.exists() else None


def persist_template(uploaded_template) -> Path:
    ensure_dirs()
    template_bytes = uploaded_template.getvalue()
    digest = image_digest(template_bytes)
    template_path = TEMPLATE_DIR / f"{digest}_{safe_name(uploaded_template.name)}"
    if not template_path.exists():
        template_path.write_bytes(template_bytes)
    return template_path


def template_store_name(template_path: Path | None) -> str:
    if not template_path or not template_path.exists():
        return DEFAULT_STORE_NAME
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(template_path, read_only=True, data_only=True)
        value = workbook.worksheets[0]["A1"].value
        workbook.close()
        return str(value).strip() if value else DEFAULT_STORE_NAME
    except Exception:
        return DEFAULT_STORE_NAME


def safe_name(filename: str) -> str:
    stem = Path(filename).stem or "image"
    suffix = Path(filename).suffix.lower() or ".jpg"
    stem = re.sub(r"[^a-zA-Z0-9._-]+", "_", stem).strip("_") or "image"
    return f"{stem[:80]}{suffix}"


def image_digest(data: bytes) -> str:
    return hashlib.sha1(data).hexdigest()[:14]


def normalize_class_list(raw: str) -> List[str]:
    return [item.strip().lower() for item in raw.split(",") if item.strip()]


def names_dict(model: YOLO) -> Dict[int, str]:
    names = getattr(model, "names", {})
    if isinstance(names, dict):
        return {int(k): str(v) for k, v in names.items()}
    return {i: str(name) for i, name in enumerate(names)}


def target_ids(model_names: Dict[int, str], target_classes: Iterable[str]) -> set[int] | None:
    targets = {name.lower() for name in target_classes}
    if not targets or "*" in targets or "all" in targets:
        return None
    return {
        class_id
        for class_id, class_name in model_names.items()
        if class_name.lower() in targets
    }


def circle_box(
    draw: ImageDraw.ImageDraw,
    box: Tuple[float, float, float, float],
    image_size: Tuple[int, int],
    large: bool = False,
) -> None:
    width, height = image_size
    x1, y1, x2, y2 = box
    pad_scale = 0.35 if large else 0.12
    min_pad = 45 if large else 10
    line_divisor = 145 if large else 220
    pad = max(min_pad, int(max(x2 - x1, y2 - y1) * pad_scale))
    line_width = max(6 if large else 4, round(max(width, height) / line_divisor))
    ellipse = [
        max(0, int(x1) - pad),
        max(0, int(y1) - pad),
        min(width, int(x2) + pad),
        min(height, int(y2) + pad),
    ]
    draw.ellipse(ellipse, outline=(255, 0, 0), width=line_width)


def box_iou(box_a: List[float], box_b: List[float]) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    inter_w, inter_h = max(0, ix2 - ix1), max(0, iy2 - iy1)
    inter_area = inter_w * inter_h
    if inter_area == 0:
        return 0

    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
    return inter_area / max(1, area_a + area_b - inter_area)


def dedupe_detections(detections: List[dict], iou_threshold: float = 0.30) -> List[dict]:
    kept: List[dict] = []
    for detection in sorted(detections, key=lambda item: item["confidence"], reverse=True):
        overlaps_existing = any(
            detection["class_name"] == existing["class_name"]
            and box_iou(detection["box_xyxy"], existing["box_xyxy"]) >= iou_threshold
            for existing in kept
        )
        if not overlaps_existing:
            kept.append(detection)
    return kept


def boxes_are_near(box_a: List[int], box_b: List[int], padding: int) -> bool:
    return not (
        box_a[2] + padding < box_b[0] - padding
        or box_b[2] + padding < box_a[0] - padding
        or box_a[3] + padding < box_b[1] - padding
        or box_b[3] + padding < box_a[1] - padding
    )


def merge_nearby_candidates(candidates: List[dict], padding: int) -> List[List[dict]]:
    clusters: List[List[dict]] = []
    for candidate in candidates:
        for cluster in clusters:
            if any(boxes_are_near(candidate["box"], item["box"], padding) for item in cluster):
                cluster.append(candidate)
                break
        else:
            clusters.append([candidate])

    changed = True
    while changed:
        changed = False
        merged: List[List[dict]] = []
        while clusters:
            cluster = clusters.pop(0)
            index = 0
            while index < len(clusters):
                if any(
                    boxes_are_near(a["box"], b["box"], padding)
                    for a in cluster
                    for b in clusters[index]
                ):
                    cluster.extend(clusters.pop(index))
                    changed = True
                else:
                    index += 1
            merged.append(cluster)
        clusters = merged

    return clusters


def small_litter_candidates(
    mask: np.ndarray,
    label: str,
    y_offset: int,
    config: dict,
    hsv: np.ndarray,
) -> List[dict]:
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8))
    if label == "colored_floor_litter":
        mask = cv2.dilate(mask, np.ones((3, 3), np.uint8), iterations=1)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    candidates = []
    for contour in contours:
        area = cv2.contourArea(contour)
        x, y, width, height = cv2.boundingRect(contour)
        aspect = max(width / height if height else 99, height / width if width else 99)

        if not (config["min_area"] <= area <= config["max_area"]):
            continue
        if width > config["max_side"] or height > config["max_side"]:
            continue
        if label == "light_floor_litter" and (aspect > 3.2 or min(width, height) < 5):
            continue
        if label == "colored_floor_litter" and aspect > 4.2:
            continue
        if label == "colored_floor_litter":
            component_mask = np.zeros(mask.shape, dtype=np.uint8)
            cv2.drawContours(component_mask, [contour], -1, 255, -1)
            component_pixels = component_mask > 0
            median_value = float(np.median(hsv[:, :, 2][component_pixels]))
            if median_value < config["colored_min_value"]:
                continue

        candidates.append(
            {
                "label": label,
                "area": area,
                "box": [x, y + y_offset, x + width, y + y_offset + height],
            }
        )
    return candidates


def detect_small_floor_litter(
    image: Image.Image,
    floor_start_ratio: float,
    sensitivity: str,
) -> List[dict]:
    config = LITTER_SENSITIVITY.get(sensitivity, LITTER_SENSITIVITY["中"])
    image_array = np.array(image)
    height, width = image_array.shape[:2]
    floor_start_y = min(height - 1, max(0, int(height * floor_start_ratio)))
    floor_region = image_array[floor_start_y:]

    hsv = cv2.cvtColor(floor_region, cv2.COLOR_RGB2HSV)
    gray = cv2.cvtColor(floor_region, cv2.COLOR_RGB2GRAY)
    background = cv2.GaussianBlur(gray, (41, 41), 0)
    bright_delta = cv2.subtract(gray, background)

    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]
    colored_mask = (
        (saturation > config["sat_threshold"]) & (value > 45)
    ).astype(np.uint8) * 255
    light_mask = (
        (bright_delta > config["bright_delta"]) & (gray > 135) & (saturation < 95)
    ).astype(np.uint8) * 255

    candidates = []
    candidates.extend(
        small_litter_candidates(colored_mask, "colored_floor_litter", floor_start_y, config, hsv)
    )
    candidates.extend(
        small_litter_candidates(light_mask, "light_floor_litter", floor_start_y, config, hsv)
    )

    detections = []
    for cluster in merge_nearby_candidates(candidates, padding=18):
        x1 = min(item["box"][0] for item in cluster)
        y1 = min(item["box"][1] for item in cluster)
        x2 = max(item["box"][2] for item in cluster)
        y2 = max(item["box"][3] for item in cluster)
        cluster_width = x2 - x1
        cluster_height = y2 - y1
        if cluster_width > config["max_cluster_side"] or cluster_height > config["max_cluster_side"]:
            continue

        area = sum(item["area"] for item in cluster)
        labels = {item["label"] for item in cluster}
        class_name = (
            "floor_litter"
            if len(labels) > 1
            else next(iter(labels))
        )
        confidence = min(0.95, 0.45 + min(area, 600) / 1200)
        detections.append(
            {
                "class_id": -1,
                "class_name": class_name,
                "confidence": round(confidence, 4),
                "source": "small_litter",
                "on_floor": True,
                "box_xyxy": [round(x1, 2), round(y1, 2), round(x2, 2), round(y2, 2)],
            }
        )

    return detections


def red_markup_mask(image_array: np.ndarray) -> np.ndarray:
    hsv = cv2.cvtColor(image_array, cv2.COLOR_RGB2HSV)
    hue, saturation, value = cv2.split(hsv)
    red = (
        ((hue < 12) | (hue > 165)) & (saturation > 65) & (value > 90)
    ).astype(np.uint8) * 255
    red = cv2.morphologyEx(red, cv2.MORPH_CLOSE, np.ones((11, 11), np.uint8), iterations=1)
    return cv2.dilate(red, np.ones((23, 23), np.uint8), iterations=2)


def watermark_overlay_mask(image_array: np.ndarray) -> np.ndarray:
    height, width = image_array.shape[:2]
    gray = cv2.cvtColor(image_array, cv2.COLOR_RGB2GRAY)
    hsv = cv2.cvtColor(image_array, cv2.COLOR_RGB2HSV)
    saturation = hsv[:, :, 1]

    bright_text = ((gray > 205) & (saturation < 80)).astype(np.uint8) * 255
    bright_text[: int(height * 0.52), :] = 0
    min_overlay_pixels = max(300, int(width * height * 0.002))
    if int((bright_text > 0).sum()) < min_overlay_pixels:
        return np.zeros((height, width), dtype=np.uint8)

    bright_text = cv2.morphologyEx(
        bright_text,
        cv2.MORPH_CLOSE,
        np.ones((23, 23), np.uint8),
        iterations=1,
    )
    return cv2.dilate(bright_text, np.ones((47, 47), np.uint8), iterations=2)


def corner_dust_candidate_mask(image_array: np.ndarray, config: dict) -> np.ndarray:
    height, width = image_array.shape[:2]
    gray = cv2.cvtColor(image_array, cv2.COLOR_RGB2GRAY)
    hsv = cv2.cvtColor(image_array, cv2.COLOR_RGB2HSV)
    saturation = hsv[:, :, 1]

    background = cv2.GaussianBlur(gray, (41, 41), 0)
    dark_delta = cv2.subtract(background, gray)
    kernel_size = max(21, int(min(width, height) / 42) // 2 * 2 + 1)
    blackhat = cv2.morphologyEx(
        gray,
        cv2.MORPH_BLACKHAT,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (kernel_size, kernel_size)),
    )
    blocked = (red_markup_mask(image_array) > 0) | (watermark_overlay_mask(image_array) > 0)

    mask = (
        (
            ((dark_delta > config["delta"]) & (gray < config["max_gray"]))
            | ((blackhat > config["delta"] + 2) & (gray < min(235, config["max_gray"] + 12)))
            | (gray < config["very_dark"])
        )
        & (saturation < 130)
        & (~blocked)
    ).astype(np.uint8)
    return cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8), iterations=1)


def detect_corner_dust(image: Image.Image, sensitivity: str) -> List[dict]:
    config = DUST_SENSITIVITY.get(sensitivity, DUST_SENSITIVITY["中"])
    image_array = np.array(image)
    height, width = image_array.shape[:2]
    mask = corner_dust_candidate_mask(image_array, config)

    window_size = max(170, min(280, int(min(width, height) * 0.22)))
    step = max(42, window_size // 4)
    integral = cv2.integral(mask)
    scores: List[int] = []

    for y in range(0, height - window_size + 1, step):
        for x in range(0, width - window_size + 1, step):
            score = int(
                integral[y + window_size, x + window_size]
                - integral[y, x + window_size]
                - integral[y + window_size, x]
                + integral[y, x]
            )
            if score > 0:
                scores.append(score)

    if not scores:
        return []

    score_threshold = max(
        config["min_window_pixels"],
        int(np.percentile(scores, config["score_percentile"])),
    )
    candidates = []
    for y in range(0, height - window_size + 1, step):
        for x in range(0, width - window_size + 1, step):
            score = int(
                integral[y + window_size, x + window_size]
                - integral[y, x + window_size]
                - integral[y + window_size, x]
                + integral[y, x]
            )
            density = score / max(1, window_size * window_size)
            if score >= score_threshold and density >= config["min_window_density"]:
                candidates.append(
                    {
                        "box": [x, y, x + window_size, y + window_size],
                        "area": score,
                    }
                )

    detections = []
    for cluster in merge_nearby_candidates(candidates, padding=step):
        x1 = min(item["box"][0] for item in cluster)
        y1 = min(item["box"][1] for item in cluster)
        x2 = max(item["box"][2] for item in cluster)
        y2 = max(item["box"][3] for item in cluster)
        region_area = max(1, (x2 - x1) * (y2 - y1))
        dirty_pixels = int(mask[y1:y2, x1:x2].sum())
        density = dirty_pixels / region_area

        if dirty_pixels < score_threshold:
            continue
        if region_area > width * height * config["max_region_ratio"] and density < 0.035:
            continue

        confidence = min(
            0.95,
            0.40
            + min(dirty_pixels / max(1, score_threshold), 4) * 0.10
            + min(len(cluster), 25) * 0.006,
        )
        detections.append(
            {
                "class_id": -2,
                "class_name": "corner_dust",
                "confidence": round(confidence, 4),
                "source": "corner_dust_rule",
                "on_floor": False,
                "large_circle": True,
                "box_xyxy": [round(x1, 2), round(y1, 2), round(x2, 2), round(y2, 2)],
            }
        )

    return sorted(detections, key=lambda item: item["confidence"], reverse=True)[:4]


def box_mask_ratio(mask: np.ndarray, box: Tuple[float, float, float, float]) -> float:
    height, width = mask.shape[:2]
    x1, y1, x2, y2 = box
    left = max(0, min(width, int(x1)))
    top = max(0, min(height, int(y1)))
    right = max(0, min(width, int(x2)))
    bottom = max(0, min(height, int(y2)))
    if right <= left or bottom <= top:
        return 0
    region = mask[top:bottom, left:right]
    return float((region > 0).sum() / max(1, region.size))


def is_in_floor_area(
    box: Tuple[float, float, float, float],
    image_height: int,
    floor_start_ratio: float,
) -> bool:
    _, y1, _, y2 = box
    center_y = (y1 + y2) / 2
    floor_start_y = image_height * floor_start_ratio
    return center_y >= floor_start_y or y2 >= floor_start_y


def detect_and_annotate(
    model: YOLO,
    image: Image.Image,
    conf: float,
    imgsz: int,
    target_classes: List[str],
    only_floor: bool,
    floor_start_ratio: float,
    detect_small_litter: bool,
    litter_sensitivity: str,
    detect_corner_dust_enabled: bool,
    corner_dust_sensitivity: str,
) -> Tuple[Image.Image, List[dict]]:
    model_names = names_dict(model)
    allowed_ids = target_ids(model_names, target_classes)
    annotated = image.copy()
    draw = ImageDraw.Draw(annotated)
    image_array = np.array(image)
    ignored_markup_mask = (
        (red_markup_mask(image_array) > 0) | (watermark_overlay_mask(image_array) > 0)
    ).astype(np.uint8)

    results = model.predict(image, conf=conf, imgsz=imgsz, verbose=False)
    detections: List[dict] = []

    for result in results:
        for box in result.boxes:
            class_id = int(box.cls.item())
            if allowed_ids is not None and class_id not in allowed_ids:
                continue

            confidence = float(box.conf.item())
            x1, y1, x2, y2 = [float(value) for value in box.xyxy[0].tolist()]
            if box_mask_ratio(ignored_markup_mask, (x1, y1, x2, y2)) > 0.35:
                continue

            class_name = model_names.get(class_id, str(class_id))
            on_floor = is_in_floor_area((x1, y1, x2, y2), image.height, floor_start_ratio)
            trained_floor_class = class_name.lower() in TRAINED_FLOOR_CLASSES
            if only_floor and not on_floor and not trained_floor_class:
                continue

            detections.append(
                {
                    "class_id": class_id,
                    "class_name": class_name,
                    "confidence": round(confidence, 4),
                    "source": "YOLO",
                    "on_floor": on_floor or trained_floor_class,
                    "box_xyxy": [round(x1, 2), round(y1, 2), round(x2, 2), round(y2, 2)],
                }
            )

    if detect_small_litter:
        for detection in detect_small_floor_litter(image, floor_start_ratio, litter_sensitivity):
            if box_mask_ratio(ignored_markup_mask, tuple(detection["box_xyxy"])) > 0.35:
                continue
            if any(box_iou(detection["box_xyxy"], item["box_xyxy"]) > 0.3 for item in detections):
                continue
            detections.append(detection)

    if detect_corner_dust_enabled:
        detections.extend(detect_corner_dust(image, corner_dust_sensitivity))

    detections = dedupe_detections(detections)
    for detection in detections:
        circle_box(
            draw,
            tuple(detection["box_xyxy"]),
            annotated.size,
            large=bool(detection.get("large_circle")),
        )
    return annotated, detections


def persist_images(uploaded_name: str, uploaded_bytes: bytes, annotated: Image.Image) -> Tuple[str, Path, Path]:
    ensure_dirs()
    digest = image_digest(uploaded_bytes)
    original_path = ORIGINAL_DIR / f"{digest}_{safe_name(uploaded_name)}"
    annotated_path = ANNOTATED_DIR / f"{digest}_red_circle.png"

    if not original_path.exists():
        original_path.write_bytes(uploaded_bytes)
    annotated.save(annotated_path)
    return digest, original_path, annotated_path


def append_review(row: dict) -> None:
    ensure_dirs()
    fieldnames = [
        "timestamp",
        "image_id",
        "human_label",
        "auto_status",
        "detection_count",
        "model_path",
        "confidence_threshold",
        "image_size",
        "original_path",
        "annotated_path",
        "detections_json",
    ]
    write_header = not LABELS_CSV.exists()
    with LABELS_CSV.open("a", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def save_review(
    human_label: str,
    image_id: str,
    original_path: Path,
    annotated_path: Path,
    detections: List[dict],
    model_path: str,
    conf: float,
    image_size: Tuple[int, int],
    auto_status: str,
) -> None:
    append_review(
        {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "image_id": image_id,
            "human_label": human_label,
            "auto_status": auto_status,
            "detection_count": len(detections),
            "model_path": model_path,
            "confidence_threshold": conf,
            "image_size": f"{image_size[0]}x{image_size[1]}",
            "original_path": str(original_path),
            "annotated_path": str(annotated_path),
            "detections_json": json.dumps(detections, ensure_ascii=False),
        }
    )


def auto_status_from_detections(detections: List[dict], only_floor: bool) -> str:
    has_floor_detection = any(item.get("on_floor") for item in detections)
    has_corner_dust = any(item.get("class_name") == "corner_dust" for item in detections)
    return (
        "发现疑似地面垃圾/角落积灰"
        if has_floor_detection and has_corner_dust
        else "地上发现疑似垃圾"
        if has_floor_detection
        else "发现疑似角落积灰"
        if has_corner_dust
        else "疑似发现垃圾"
        if detections
        else "未发现地上垃圾"
        if only_floor
        else "未发现疑似垃圾"
    )


def summarize_detection_list(detections: List[dict]) -> str:
    counts: Dict[str, int] = {}
    for detection in detections:
        class_name = str(detection.get("class_name", "unknown"))
        counts[class_name] = counts.get(class_name, 0) + 1
    return ", ".join(f"{name} x{count}" for name, count in sorted(counts.items()))


def image_path_to_data_url(image_path: Path, max_size: Tuple[int, int] = (220, 150)) -> str:
    with Image.open(image_path).convert("RGB") as image:
        image.thumbnail(max_size)
        buffer = BytesIO()
        image.save(buffer, format="JPEG", quality=82)
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{encoded}"


def detection_detail_rows(detections: List[dict]) -> List[dict]:
    return [
        {
            "来源": item.get("source", "YOLO"),
            "类别": item["class_name"],
            "置信度": item["confidence"],
            "区域": "地上"
            if item.get("on_floor")
            else "边角"
            if item.get("class_name") == "corner_dust"
            else "非地面",
            "位置": item["box_xyxy"],
        }
        for item in detections
    ]


def batch_result_to_table_row(result: dict) -> dict:
    score = parse_deduction_score(result.get("deduction_score"))
    return {
        "红圈图": image_path_to_data_url(Path(result["annotated_path"])),
        "人工判定": result.get("human_label") or "待审核",
        "照片名称": result["filename"],
        "自动判断": result["auto_status"],
        "扣分项": result.get("deduction_project", ""),
        "扣分内容": result.get("deduction_content", ""),
        "扣分": format_score_value(score) if score is not None else None,
        "红圈数量": result["detection_count"],
        "检测类别": summarize_detection_list(result["detections"]),
        "图片ID": result["image_id"],
    }


def table_records(edited_table) -> List[dict]:
    if hasattr(edited_table, "to_dict"):
        return edited_table.to_dict("records")
    return list(edited_table)


def merged_cell_value(worksheet, row: int, column: int):
    cell = worksheet.cell(row, column)
    if cell.value is not None:
        return cell.value
    coordinate = cell.coordinate
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


def normalize_text(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def clean_editor_text(value) -> str:
    return str(value or "").strip()


def parse_deduction_score(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        match = re.search(r"-?\d+(?:\.\d+)?", value)
        if not match:
            return None
        raw_score = match.group()
    else:
        raw_score = value

    try:
        score = float(raw_score)
    except (TypeError, ValueError):
        return None
    if score != score:
        return None
    return -abs(score) if score > 0 else score


def format_score_value(score: float | None):
    if score is None:
        return ""
    return int(score) if float(score).is_integer() else score


def normalized_row_number(value) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def extract_inspection_items(template_path: Path | None) -> List[dict]:
    if not template_path or not template_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return []

    workbook = load_workbook(template_path, data_only=False)
    worksheet = workbook.worksheets[0]
    items = []
    for row in range(1, worksheet.max_row + 1):
        score_cell = None
        score_value = None
        for column in range(1, min(worksheet.max_column, 20) + 1):
            value = merged_cell_value(worksheet, row, column)
            if isinstance(value, (int, float)) and value < 0:
                score_cell = worksheet.cell(row, column).coordinate
                score_value = value
                break
        if score_cell is None:
            continue

        category = merged_cell_value(worksheet, row, 1) or ""
        project = ""
        for column in range(2, 6):
            value = merged_cell_value(worksheet, row, column)
            if value:
                project = str(value)
                break
        content = ""
        for column in range(6, 8):
            value = merged_cell_value(worksheet, row, column)
            if value:
                content = str(value)
                break

        items.append(
            {
                "row": row,
                "score_cell": score_cell,
                "category": str(category),
                "project": project,
                "content": content,
                "score": score_value,
                "text": normalize_text(f"{category} {project} {content}"),
            }
        )
    workbook.close()
    return items


def match_deduction_item(result: dict, inspection_items: List[dict]) -> dict | None:
    detections = result.get("detections", [])
    if not detections or not inspection_items:
        return None

    class_names = {str(item.get("class_name", "")).lower() for item in detections}
    source_text = normalize_text(
        " ".join(
            [
                result.get("filename", ""),
                result.get("auto_status", ""),
                summarize_detection_list(detections),
            ]
        )
    )

    dirty_classes = {
        "corner_dust",
        "dust",
        "stain",
        "dirty",
        "residue",
        "mold",
        "oil",
        "grease",
        "floor_litter",
        "light_floor_litter",
        "colored_floor_litter",
        "paper_scrap",
        "food residue",
    }
    scored_items = []
    for item in inspection_items:
        text = item["text"]
        score = 0
        if any(keyword in source_text for keyword in ["门头", "蜘蛛", "外观"]):
            score += 10 if any(keyword in text for keyword in ["门头", "蜘蛛", "外观"]) else 0
        if "展架" in source_text:
            score += 10 if "展架" in text else 0
        if "货架" in source_text or "私人物品" in source_text:
            score += 10 if any(keyword in text for keyword in ["货架", "私人物品", "混放"]) else 0
        if class_names & dirty_classes:
            score += 8 if any(keyword in text for keyword in ["店铺内部", "桌面", "底坐", "灯带", "蚊虫", "清理", "擦拭", "干净", "玻璃"]) else 0
            score += 3 if any(keyword in text for keyword in ["卫生", "内部", "店铺"]) else 0
        if "corner_dust" in class_names:
            score += 4 if any(keyword in text for keyword in ["灯带", "蚊虫", "蜘蛛", "清理", "角"]) else 0
        scored_items.append((score, item))

    scored_items.sort(key=lambda pair: (pair[0], abs(pair[1]["score"])), reverse=True)
    best_score, best_item = scored_items[0]
    if best_score <= 0:
        for item in inspection_items:
            if "店铺内部" in item["text"]:
                return item
        return inspection_items[0]
    return best_item


def image_id_from_path(path: Path) -> str:
    return path.name.split("_", 1)[0]


def summarize_detections(raw_json: str) -> str:
    if not raw_json:
        return ""
    try:
        detections = json.loads(raw_json)
    except json.JSONDecodeError:
        return ""

    counts: Dict[str, int] = {}
    for detection in detections:
        class_name = str(detection.get("class_name", "unknown"))
        counts[class_name] = counts.get(class_name, 0) + 1
    return ", ".join(f"{name} x{count}" for name, count in sorted(counts.items()))


def latest_review_rows() -> Dict[str, dict]:
    if not LABELS_CSV.exists():
        return {}

    reviews: Dict[str, dict] = {}
    with LABELS_CSV.open("r", newline="", encoding="utf-8") as csvfile:
        for row in csv.DictReader(csvfile):
            image_id = row.get("image_id", "")
            if image_id:
                reviews[image_id] = row
    return reviews


def collect_excel_rows() -> List[dict]:
    originals = {
        image_id_from_path(path): path
        for path in ORIGINAL_DIR.glob("*")
        if path.is_file()
    }
    annotated = {
        image_id_from_path(path): path
        for path in ANNOTATED_DIR.glob("*_red_circle.png")
        if path.is_file()
    }
    reviews = latest_review_rows()
    image_ids = sorted(
        set(originals) | set(annotated) | set(reviews),
        key=lambda image_id: max(
            [
                path.stat().st_mtime
                for path in [originals.get(image_id), annotated.get(image_id)]
                if path and path.exists()
            ]
            or [0]
        ),
        reverse=True,
    )

    rows = []
    for image_id in image_ids:
        review = reviews.get(image_id, {})
        original_path = originals.get(image_id)
        annotated_path = annotated.get(image_id)
        rows.append(
            {
                "image_id": image_id,
                "filename": original_path.name.split("_", 1)[1] if original_path and "_" in original_path.name else "",
                "human_label": review.get("human_label", ""),
                "auto_status": review.get("auto_status", ""),
                "detection_count": review.get("detection_count", ""),
                "detection_summary": summarize_detections(review.get("detections_json", "")),
                "timestamp": review.get("timestamp", ""),
                "original_path": original_path,
                "annotated_path": annotated_path,
            }
        )
    return rows


def add_excel_image(worksheet, image_path: Path | None, cell: str, max_width: int = 128, max_height: int = 96) -> None:
    if not image_path or not image_path.exists():
        return

    from openpyxl.drawing.image import Image as ExcelImage

    with Image.open(image_path) as image:
        width, height = image.size

    scale = min(max_width / max(1, width), max_height / max(1, height), 1)
    excel_image = ExcelImage(str(image_path))
    excel_image.width = int(width * scale)
    excel_image.height = int(height * scale)
    worksheet.add_image(excel_image, cell)


def photo_cell_for_index(index: int) -> str:
    row_group = index // len(PHOTO_GRID_COLUMNS)
    column = PHOTO_GRID_COLUMNS[index % len(PHOTO_GRID_COLUMNS)]
    row = PHOTO_GRID_START_ROW + row_group * PHOTO_GRID_ROW_STEP
    return f"{column}{row}"


def prepare_excel_photo(source_path: Path, output_dir: Path, index: int) -> Path:
    image = Image.open(source_path).convert("RGB")
    image = ImageOps.contain(image, PHOTO_CANVAS_SIZE)
    canvas = Image.new("RGB", PHOTO_CANVAS_SIZE, "white")
    left = (PHOTO_CANVAS_SIZE[0] - image.width) // 2
    top = (PHOTO_CANVAS_SIZE[1] - image.height) // 2
    canvas.paste(image, (left, top))
    output_path = output_dir / f"photo_{index:03d}.jpg"
    canvas.save(output_path, format="JPEG", quality=90)
    return output_path


def format_photo_grid(worksheet, photo_count: int) -> None:
    row_groups = max(1, (photo_count + len(PHOTO_GRID_COLUMNS) - 1) // len(PHOTO_GRID_COLUMNS))
    for row_group in range(row_groups):
        start_row = PHOTO_GRID_START_ROW + row_group * PHOTO_GRID_ROW_STEP
        for row in range(start_row, start_row + PHOTO_GRID_ROW_STEP):
            worksheet.row_dimensions[row].height = 20


def parse_total_score(value) -> int | None:
    match = re.search(r"-?\d+", str(value or ""))
    return int(match.group()) if match else None


def find_total_cell(worksheet):
    total_cell = None
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "总分" in cell.value:
                return cell
    return total_cell


def row_text(worksheet, row: int, max_column: int = 12) -> str:
    values = [
        str(merged_cell_value(worksheet, row, column) or "")
        for column in range(1, min(worksheet.max_column, max_column) + 1)
    ]
    return " ".join(value for value in values if value)


def inspection_table_end_row(worksheet) -> int:
    search_limit = min(max(worksheet.max_row, PHOTO_GRID_START_ROW), PHOTO_GRID_START_ROW + 12)
    for row in range(1, search_limit + 1):
        text = row_text(worksheet, row)
        if "温馨提示" in text or "总分" in text:
            return row
    return PHOTO_GRID_START_ROW


def is_empty_cell_value(value) -> bool:
    return str(value or "").strip() == ""


def manual_deduction_target_rows(worksheet, needed_count: int) -> List[int]:
    if needed_count <= 0:
        return []

    end_row = max(4, min(inspection_table_end_row(worksheet), PHOTO_GRID_START_ROW))
    blocked_terms = ["巡检项目", "巡检内容", "考核", "食品安全合规项", "服务合规项", "温馨提示", "总分"]
    target_rows = []

    for row in range(3, end_row):
        text = row_text(worksheet, row)
        if any(term in text for term in blocked_terms):
            continue
        project = merged_cell_value(worksheet, row, 2)
        content = merged_cell_value(worksheet, row, 6)
        score = merged_cell_value(worksheet, row, 8)
        if is_empty_cell_value(project) and is_empty_cell_value(content) and is_empty_cell_value(score):
            target_rows.append(row)
        if len(target_rows) >= needed_count:
            return target_rows

    return target_rows


def merged_cell_anchor(worksheet, row: int, column: int) -> Tuple[int, int]:
    coordinate = worksheet.cell(row, column).coordinate
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range.min_row, merged_range.min_col
    return row, column


def write_merged_cell(worksheet, row: int, column: int, value) -> None:
    anchor_row, anchor_column = merged_cell_anchor(worksheet, row, column)
    worksheet.cell(anchor_row, anchor_column).value = value


def template_base_total(worksheet, inspection_items: List[dict]) -> int:
    total_cell = find_total_cell(worksheet)
    original_total = parse_total_score(total_cell.value) if total_cell else None
    original_negative_total = sum(float(item["score"]) for item in inspection_items)
    return (
        int(original_total - original_negative_total)
        if original_total is not None
        else 200
    )


def apply_deductions_to_sheet(worksheet, inspection_items: List[dict], rows: List[dict], base_total: int) -> None:
    total_cell = find_total_cell(worksheet)
    item_by_row = {item["row"]: item for item in inspection_items}
    for item in inspection_items:
        worksheet[item["score_cell"]] = None

    matched_rows: Dict[int, float] = {}
    manual_rows = []
    for row in rows:
        score = parse_deduction_score(row.get("deduction_score"))
        if score is None:
            continue

        item_row = normalized_row_number(row.get("deduction_row"))
        if item_row in item_by_row:
            item = item_by_row[item_row]
            matched_rows[item_row] = score
            continue

        project = clean_editor_text(row.get("deduction_project"))
        content = clean_editor_text(row.get("deduction_content"))
        manual_rows.append(
            {
                "project": project or "店铺卫生",
                "content": content or clean_editor_text(row.get("auto_status")) or "手动扣分",
                "score": score,
            }
        )

    for item_row, score in matched_rows.items():
        item = item_by_row[item_row]
        worksheet[item["score_cell"]] = format_score_value(score)

    target_rows = manual_deduction_target_rows(worksheet, len(manual_rows))
    if len(target_rows) < len(manual_rows) and target_rows:
        overflow = manual_rows[len(target_rows) - 1 :]
        manual_rows = manual_rows[: len(target_rows) - 1] + [
            {
                "project": "其他扣分",
                "content": "；".join(item["content"] for item in overflow if item["content"]),
                "score": sum(float(item["score"]) for item in overflow),
            }
        ]
    else:
        manual_rows = manual_rows[: len(target_rows)]

    for target_row, item in zip(target_rows, manual_rows):
        write_merged_cell(worksheet, target_row, 2, item["project"])
        write_merged_cell(worksheet, target_row, 6, item["content"])
        write_merged_cell(worksheet, target_row, 8, format_score_value(item["score"]))

    if total_cell is not None:
        total_score = base_total + sum(matched_rows.values()) + sum(float(item["score"]) for item in manual_rows)
        total_cell.value = f"总分： {int(total_score) if float(total_score).is_integer() else total_score}"


def write_table_excel_rows(rows: List[dict], output_dir: Path, filename_prefix: str) -> Tuple[bytes, Path, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise RuntimeError("缺少 openpyxl，请先运行：pip install -r requirements.txt") from exc

    if not rows:
        raise ValueError("还没有可导出的照片。请先上传并识别图片。")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "卫生复合"
    headers = [
        "序号",
        "图片ID",
        "照片名称",
        "人工判定",
        "自动判断",
        "扣分项",
        "扣分内容",
        "扣分",
        "红圈数量",
        "检测类别",
        "记录时间",
        "原图",
        "红圈图",
        "原图路径",
        "红圈图路径",
    ]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="E6F0F8")
    header_font = Font(bold=True, color="1F2937")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        original_path = row["original_path"]
        annotated_path = row["annotated_path"]
        worksheet.append(
            [
                row_index - 1,
                row["image_id"],
                row["filename"],
                row["human_label"],
                row["auto_status"],
                row.get("deduction_project", ""),
                row.get("deduction_content", ""),
                format_score_value(parse_deduction_score(row.get("deduction_score"))),
                row["detection_count"],
                row["detection_summary"],
                row["timestamp"],
                "",
                "",
                str(original_path) if original_path else "",
                str(annotated_path) if annotated_path else "",
            ]
        )
        worksheet.row_dimensions[row_index].height = 76
        add_excel_image(worksheet, original_path, f"L{row_index}")
        add_excel_image(worksheet, annotated_path, f"M{row_index}")

    widths = [8, 18, 28, 12, 24, 16, 42, 10, 10, 28, 22, 18, 18, 58, 58]
    for column_index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{filename_prefix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    buffer = BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()
    output_path.write_bytes(data)
    return data, output_path, len(rows)


def write_template_excel_rows(
    rows: List[dict],
    output_dir: Path,
    filename_prefix: str,
    store_name: str,
    template_path: Path | None,
) -> Tuple[bytes, Path, int]:
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as ExcelImage
    except ModuleNotFoundError as exc:
        raise RuntimeError("缺少 openpyxl，请先运行：pip install -r requirements.txt") from exc

    if not rows:
        raise ValueError("还没有可导出的照片。请先上传并识别图片。")
    if not template_path or not template_path.exists():
        return write_table_excel_rows(rows, output_dir, filename_prefix)

    workbook = load_workbook(template_path)
    template_sheet = workbook.worksheets[0]
    inspection_items = extract_inspection_items(template_path)
    base_total = template_base_total(template_sheet, inspection_items)

    template_sheet._images = []
    for sheet in list(workbook.worksheets[1:]):
        workbook.remove(sheet)
    template_sheet.title = "_导出模板"

    rows_per_sheet = 12
    with tempfile.TemporaryDirectory() as temp_dir:
        for chunk_index, start in enumerate(range(0, len(rows), rows_per_sheet)):
            chunk = rows[start : start + rows_per_sheet]
            worksheet = workbook.copy_worksheet(template_sheet)
            worksheet.title = "卫生复合" if chunk_index == 0 else f"卫生复合{chunk_index + 1}"
            worksheet._images = []

            worksheet["A1"] = store_name or DEFAULT_STORE_NAME
            apply_deductions_to_sheet(worksheet, inspection_items, chunk, base_total)
            worksheet._images = []
            format_photo_grid(worksheet, len(chunk))
            temp_path = Path(temp_dir)
            for image_index, row in enumerate(chunk):
                annotated_path = row.get("annotated_path")
                if not annotated_path:
                    continue
                image_path = Path(annotated_path)
                if not image_path.exists():
                    continue

                prepared_path = prepare_excel_photo(image_path, temp_path, image_index)
                excel_image = ExcelImage(str(prepared_path))
                excel_image.width = PHOTO_DISPLAY_SIZE[0]
                excel_image.height = PHOTO_DISPLAY_SIZE[1]
                excel_image.anchor = photo_cell_for_index(image_index)
                worksheet.add_image(excel_image)

        workbook.remove(template_sheet)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{filename_prefix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        buffer = BytesIO()
        workbook.save(buffer)
        data = buffer.getvalue()
        output_path.write_bytes(data)
    return data, output_path, len(rows)


def write_excel_rows(
    rows: List[dict],
    output_dir: Path,
    filename_prefix: str,
    store_name: str,
    template_path: Path | None,
) -> Tuple[bytes, Path, int]:
    return write_template_excel_rows(rows, output_dir, filename_prefix, store_name, template_path)


def build_excel_report(store_name: str, template_path: Path | None) -> Tuple[bytes, Path, int]:
    return write_excel_rows(collect_excel_rows(), EXCEL_REPORT_DIR, "卫生复合照片汇总", store_name, template_path)


def current_results_to_excel_rows(results: List[dict], label_by_id: Dict[str, str]) -> List[dict]:
    rows = []
    now = datetime.now().isoformat(timespec="seconds")
    for result in results:
        label = label_by_id.get(result["image_id"], result.get("human_label", "待审核"))
        score = parse_deduction_score(result.get("deduction_score"))
        rows.append(
            {
                "image_id": result["image_id"],
                "filename": result["filename"],
                "human_label": label,
                "auto_status": result["auto_status"],
                "deduction_project": result.get("deduction_project", ""),
                "deduction_content": result.get("deduction_content", ""),
                "deduction_score": format_score_value(score) if score is not None else "",
                "deduction_row": result.get("deduction_row"),
                "detection_count": result["detection_count"],
                "detection_summary": summarize_detection_list(result["detections"]),
                "timestamp": now,
                "original_path": Path(result["original_path"]),
                "annotated_path": Path(result["annotated_path"]),
            }
        )
    return rows


def sync_results_from_editor(results: List[dict], edited_rows: List[dict], inspection_items: List[dict]) -> List[dict]:
    row_by_id = {str(row.get("图片ID", "")): row for row in edited_rows}
    item_by_row = {item["row"]: item for item in inspection_items}
    synced_results = []

    for result in results:
        row = row_by_id.get(result["image_id"])
        if not row:
            synced_results.append(result)
            continue

        result = dict(result)
        label = clean_editor_text(row.get("人工判定")) or "待审核"
        result["human_label"] = label if label in {"待审核", "合格", "不合格"} else "待审核"

        project = clean_editor_text(row.get("扣分项", row.get("自动扣分项", "")))
        content = clean_editor_text(row.get("扣分内容"))
        score = parse_deduction_score(row.get("扣分"))
        result["deduction_project"] = project
        result["deduction_content"] = content
        result["deduction_score"] = format_score_value(score) if score is not None else ""

        item_row = normalized_row_number(result.get("deduction_row"))
        item = item_by_row.get(item_row)
        matches_template_item = bool(
            item
            and normalize_text(project) == normalize_text(item["project"])
            and normalize_text(content) == normalize_text(item["content"])
        )
        if matches_template_item:
            result["deduction_row"] = item_row
        else:
            result.pop("deduction_row", None)

        synced_results.append(result)

    return synced_results


def build_current_excel_report(
    results: List[dict],
    label_by_id: Dict[str, str],
    output_dir: Path,
    store_name: str,
    template_path: Path | None,
    filename_prefix: str = "卫生复合当前汇总",
) -> Tuple[bytes, Path, int]:
    return write_excel_rows(
        current_results_to_excel_rows(results, label_by_id),
        output_dir,
        filename_prefix,
        store_name,
        template_path,
    )


def main() -> None:
    st.set_page_config(page_title="卫生复合", layout="wide")
    st.title("卫生复合")

    with st.sidebar:
        uploaded_template = st.file_uploader("上传巡检表档案", type=["xlsx"])
        template_path = persist_template(uploaded_template) if uploaded_template else default_template_path()
        store_name = template_store_name(template_path)
        inspection_items = extract_inspection_items(template_path)
        if template_path:
            st.caption(f"已读取：{store_name}，扣分项 {len(inspection_items)} 条")
            if not inspection_items:
                st.caption("档案里没有读取到负分扣分项，可在汇总表中手动填写。")
        else:
            st.caption("未上传巡检表档案，将使用普通汇总格式导出")
        model_path = st.text_input("YOLO 模型", value=DEFAULT_MODEL_PATH)
        conf = st.slider("置信度", min_value=0.01, max_value=0.90, value=DEFAULT_CONFIDENCE, step=0.01)
        imgsz = st.select_slider("图片尺寸", options=[416, 512, 640, 768, 960], value=DEFAULT_IMAGE_SIZE)
        only_floor = st.checkbox("YOLO 只保留地面目标", value=True)
        floor_start_percent = st.slider("地面区域起始高度", 0, 80, DEFAULT_FLOOR_START_PERCENT, 5)
        detect_small_litter = st.checkbox("补充检测小碎屑", value=not HAS_TRAINED_MODEL)
        litter_sensitivity = st.select_slider("小碎屑敏感度", options=["低", "中", "高"], value="中")
        detect_corner_dust_enabled = st.checkbox("大圈标记角落积灰", value=True)
        corner_dust_sensitivity = st.select_slider("角落积灰敏感度", options=["低", "中", "高"], value="中")
        target_class_text = st.text_area("检测类别", value=DEFAULT_TARGET_CLASSES, height=130)
    uploaded_files = st.file_uploader(
        "上传卫生照片（可多选）",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
    )

    file_signatures = [
        {
            "name": uploaded.name,
            "digest": image_digest(uploaded.getvalue()),
        }
        for uploaded in uploaded_files
    ]
    settings_signature = hashlib.sha1(
        json.dumps(
            {
                "files": file_signatures,
                "template_path": str(template_path) if template_path else "",
                "template_mtime": template_path.stat().st_mtime if template_path and template_path.exists() else 0,
                "model_path": model_path,
                "conf": conf,
                "imgsz": imgsz,
                "only_floor": only_floor,
                "floor_start_percent": floor_start_percent,
                "detect_small_litter": detect_small_litter,
                "litter_sensitivity": litter_sensitivity,
                "detect_corner_dust_enabled": detect_corner_dust_enabled,
                "corner_dust_sensitivity": corner_dust_sensitivity,
                "target_class_text": target_class_text,
            },
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()

    reload_col, clear_col = st.columns([1, 5])
    force_redetect = reload_col.button(
        "重新检测",
        use_container_width=True,
        disabled=not uploaded_files,
    )
    if clear_col.button("清空当前汇总", use_container_width=True):
        st.session_state.pop("batch_results", None)
        st.session_state.pop("batch_signature", None)
        st.session_state["review_editor_version"] = st.session_state.get("review_editor_version", 0) + 1
        st.rerun()

    should_detect = bool(uploaded_files) and (
        force_redetect or st.session_state.get("batch_signature") != settings_signature
    )
    if should_detect:
        results = []
        errors = []
        previous_reviews = latest_review_rows()
        progress = st.progress(0, text="正在准备模型...")
        status_box = st.empty()

        try:
            model = load_model(model_path)
        except Exception as exc:
            progress.empty()
            st.error(f"模型加载失败：{exc}")
            st.stop()

        for index, uploaded in enumerate(uploaded_files, start=1):
            status_box.write(f"正在识别：{uploaded.name}（{index}/{len(uploaded_files)}）")
            try:
                uploaded_bytes = uploaded.getvalue()
                image = Image.open(BytesIO(uploaded_bytes)).convert("RGB")
                annotated, detections = detect_and_annotate(
                    model=model,
                    image=image,
                    conf=conf,
                    imgsz=imgsz,
                    target_classes=normalize_class_list(target_class_text),
                    only_floor=only_floor,
                    floor_start_ratio=floor_start_percent / 100,
                    detect_small_litter=detect_small_litter,
                    litter_sensitivity=litter_sensitivity,
                    detect_corner_dust_enabled=detect_corner_dust_enabled,
                    corner_dust_sensitivity=corner_dust_sensitivity,
                )
                image_id, original_path, annotated_path = persist_images(
                    uploaded.name,
                    uploaded_bytes,
                    annotated,
                )
                auto_status = auto_status_from_detections(detections, only_floor)
                previous_label = previous_reviews.get(image_id, {}).get("human_label") or "待审核"
                result = {
                    "image_id": image_id,
                    "filename": uploaded.name,
                    "original_path": str(original_path),
                    "annotated_path": str(annotated_path),
                    "detections": detections,
                    "detection_count": len(detections),
                    "auto_status": auto_status,
                    "human_label": previous_label if previous_label in {"合格", "不合格"} else "待审核",
                    "image_size": image.size,
                    "model_path": model_path,
                    "confidence_threshold": conf,
                }
                deduction_item = match_deduction_item(result, inspection_items)
                if deduction_item:
                    result.update(
                        {
                            "deduction_project": deduction_item["project"],
                            "deduction_content": deduction_item["content"],
                            "deduction_score": deduction_item["score"],
                            "deduction_row": deduction_item["row"],
                        }
                    )
                results.append(
                    result
                )
            except Exception as exc:
                errors.append(f"{uploaded.name}：{exc}")
            progress.progress(index / len(uploaded_files), text=f"已完成 {index}/{len(uploaded_files)}")

        progress.empty()
        status_box.empty()
        if errors:
            st.warning("部分图片识别失败：" + "；".join(errors))
        st.session_state["batch_results"] = results
        st.session_state["batch_signature"] = settings_signature
        st.session_state["review_editor_version"] = st.session_state.get("review_editor_version", 0) + 1

    results = st.session_state.get("batch_results", [])
    if not results:
        st.info("请上传一张或多张照片，识别后会在这里生成汇总表格。")
        return
    for result in results:
        has_manual_deduction = any(
            clean_editor_text(result.get(key))
            for key in ["deduction_project", "deduction_content", "deduction_score"]
        )
        if result.get("deduction_row") or has_manual_deduction or not result.get("detections"):
            continue
        deduction_item = match_deduction_item(result, inspection_items)
        if deduction_item:
            result.update(
                {
                    "deduction_project": deduction_item["project"],
                    "deduction_content": deduction_item["content"],
                    "deduction_score": deduction_item["score"],
                    "deduction_row": deduction_item["row"],
                }
            )
    st.session_state["batch_results"] = results

    st.subheader("检测汇总")
    total_count = len(results)
    problem_count = sum(1 for result in results if result["detection_count"] > 0)
    reviewed_count = sum(1 for result in results if result.get("human_label") in {"合格", "不合格"})
    metric_col_1, metric_col_2, metric_col_3 = st.columns(3)
    metric_col_1.metric("照片数量", total_count)
    metric_col_2.metric("疑似问题照片", problem_count)
    metric_col_3.metric("已审核", reviewed_count)

    action_col_1, action_col_2, action_col_3 = st.columns(3)
    if action_col_1.button("全部标为合格", use_container_width=True):
        for result in results:
            result["human_label"] = "合格"
        st.session_state["batch_results"] = results
        st.session_state["review_editor_version"] = st.session_state.get("review_editor_version", 0) + 1
        st.rerun()
    if action_col_2.button("全部标为不合格", use_container_width=True):
        for result in results:
            result["human_label"] = "不合格"
        st.session_state["batch_results"] = results
        st.session_state["review_editor_version"] = st.session_state.get("review_editor_version", 0) + 1
        st.rerun()
    if action_col_3.button("全部恢复待审核", use_container_width=True):
        for result in results:
            result["human_label"] = "待审核"
        st.session_state["batch_results"] = results
        st.session_state["review_editor_version"] = st.session_state.get("review_editor_version", 0) + 1
        st.rerun()

    summary_rows = [batch_result_to_table_row(result) for result in results]
    editor_key = (
        f"review_editor_{st.session_state.get('batch_signature', 'manual')}_"
        f"{st.session_state.get('review_editor_version', 0)}"
    )
    edited_table = st.data_editor(
        summary_rows,
        key=editor_key,
        hide_index=True,
        use_container_width=True,
        row_height=126,
        height=min(720, 88 + len(summary_rows) * 126),
        disabled=["红圈图", "照片名称", "自动判断", "红圈数量", "检测类别", "图片ID"],
        column_config={
            "红圈图": st.column_config.ImageColumn("红圈图", width="medium"),
            "人工判定": st.column_config.SelectboxColumn(
                "人工判定",
                options=["待审核", "合格", "不合格"],
                required=True,
                width="small",
            ),
            "照片名称": st.column_config.TextColumn("照片名称", width="medium"),
            "自动判断": st.column_config.TextColumn("自动判断", width="medium"),
            "扣分项": st.column_config.TextColumn("扣分项", width="medium"),
            "扣分内容": st.column_config.TextColumn("扣分内容", width="large"),
            "扣分": st.column_config.NumberColumn("扣分", width="small", step=1),
            "红圈数量": st.column_config.NumberColumn("红圈数量", width="small"),
            "检测类别": st.column_config.TextColumn("检测类别", width="large"),
            "图片ID": st.column_config.TextColumn("图片ID", width="small"),
        },
    )

    edited_records = table_records(edited_table)
    results = sync_results_from_editor(results, edited_records, inspection_items)
    st.session_state["batch_results"] = results
    label_by_id = {
        result["image_id"]: result.get("human_label", "待审核")
        for result in results
    }

    save_col, desktop_col, excel_col, csv_col = st.columns(4)
    if save_col.button("保存审核结果", type="primary", use_container_width=True):
        saved_count = 0
        for result in results:
            label = label_by_id.get(result["image_id"], "待审核")
            result["human_label"] = label
            if label not in {"合格", "不合格"}:
                continue
            save_review(
                label,
                result["image_id"],
                Path(result["original_path"]),
                Path(result["annotated_path"]),
                result["detections"],
                result["model_path"],
                result["confidence_threshold"],
                result["image_size"],
                result["auto_status"],
            )
            saved_count += 1
        st.session_state["batch_results"] = results
        st.success(f"已保存 {saved_count} 张照片的审核结果")

    if desktop_col.button("导出当前表格到桌面", use_container_width=True):
        try:
            excel_data, excel_path, row_count = build_current_excel_report(
                results,
                label_by_id,
                DESKTOP_DIR,
                store_name,
                template_path,
            )
            st.session_state["excel_report"] = {
                "data": excel_data,
                "path": str(excel_path),
                "row_count": row_count,
            }
            st.success(f"已导出到桌面：{excel_path}")
        except Exception as exc:
            st.error(f"导出到桌面失败：{exc}")

    if excel_col.button("生成当前表格 Excel", use_container_width=True):
        try:
            excel_data, excel_path, row_count = build_current_excel_report(
                results,
                label_by_id,
                EXCEL_REPORT_DIR,
                store_name,
                template_path,
            )
            st.session_state["excel_report"] = {
                "data": excel_data,
                "path": str(excel_path),
                "row_count": row_count,
            }
            st.success(f"已生成 {row_count} 张照片的 Excel")
        except Exception as exc:
            st.error(f"生成 Excel 失败：{exc}")

    if LABELS_CSV.exists():
        csv_col.download_button(
            "下载审核记录 CSV",
            data=LABELS_CSV.read_bytes(),
            file_name=LABELS_CSV.name,
            mime="text/csv",
            use_container_width=True,
        )

    excel_report = st.session_state.get("excel_report")
    if excel_report:
        st.download_button(
            "下载已生成 Excel",
            data=excel_report["data"],
            file_name=Path(excel_report["path"]).name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.subheader("逐张查看")
    for index, result in enumerate(results, start=1):
        label = result.get("human_label", "待审核")
        with st.expander(
            f"{index}. {result['filename']} | {label} | {result['auto_status']} | 红圈 {result['detection_count']} 个",
            expanded=total_count == 1,
        ):
            image_col_1, image_col_2 = st.columns(2)
            image_col_1.image(result["original_path"], caption="原图", use_container_width=True)
            image_col_2.image(result["annotated_path"], caption="红圈图", use_container_width=True)
            if result["detections"]:
                st.dataframe(
                    detection_detail_rows(result["detections"]),
                    hide_index=True,
                    use_container_width=True,
                )
            else:
                st.write("未检测到疑似问题。")
            st.download_button(
                "下载这张红圈图",
                data=Path(result["annotated_path"]).read_bytes(),
                file_name=Path(result["annotated_path"]).name,
                mime="image/png",
                use_container_width=True,
                key=f"download_annotated_{result['image_id']}",
            )


if __name__ == "__main__":
    main()
