#!/usr/bin/env python3
"""Generate idempotent SQL for the historical monthly loss workbook.

The workbook stores one row per store and one column per item. Monthly monetary
totals are item-level, so the importer preserves them in archive tables instead
of inventing daily records or allocating money to stores.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import openpyxl


SHEET_MONTHS = {
    "8月份": "2025-08",
    "10月份": "2025-10",
    "11月份": "2025-11",
    "12月份": "2025-12",
    "1月": "2026-01",
    "2月": "2026-02",
    "3月份": "2026-03",
    "4月份": "2026-04",
    "5月份": "2026-05",
    "6月份": "2026-06",
}

STORE_ALIASES = {
    "长大": "rg8",
    "万达二": "rg4",
    "万达三": "rg5",
    "火车站": "rg9",
    "花台": "rg11",
    "新天地": "rg12",
    "吾悦1": "rg6",
    "吾悦2": "rg7",
    "美佳华": "rg2",
    "荆州之星": "rg1",
    "人信汇": "rg3",
    "荆江之星": "rg13",
    "宜昌万达": "rg10",
    "新南门": "hist-xnm",
    "宜昌CBD": "hist-yccbd",
    "公安": "hist-ga",
    "松滋": "hist-sz",
    "江陵": "hist-jl",
}

HISTORICAL_STORES = {
    "hist-xnm": "新南门店",
    "hist-yccbd": "宜昌CBD店",
    "hist-ga": "公安店",
    "hist-sz": "松滋店",
    "hist-jl": "江陵店",
}

ITEM_CODES = {
    "芒果": "FRUIT_CHECK_003",
    "青芒": "FRUIT_CHECK_004",
    "火龙果": "FRUIT_CHECK_005",
    "橙子": "FRUIT_CHECK_006",
    "凤梨": "FRUIT_CHECK_007",
    "秋月梨": "FRUIT_CHECK_008",
    "苹果": "FRUIT_CHECK_009",
    "百香果": "FRUIT_CHECK_010",
    "芭乐": "FRUIT_CHECK_011",
    "杨梅": "FRUIT_CHECK_012",
    "荔枝": "FRUIT_CHECK_013",
    "黄皮": "FRUIT_CHECK_015",
    "羽衣甘蓝": "FRUIT_CHECK_018",
    "葡萄": "FRUIT_CHECK_019",
    "新鲜牛油果": "FRUIT_CHECK_002",
}

SUMMARY_LABELS = {
    "合计（总量）", "总量", "价格", "总计损耗金额", "厂商赔付金额", "店铺承担"
}
MONEY_SCALE = Decimal("0.000001")
QUANTITY_SCALE = Decimal("0.0001")
ZERO = Decimal("0")


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def q(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def number(value: Decimal | None, scale: Decimal = MONEY_SCALE) -> str:
    if value is None:
        return "null"
    return str(value.quantize(scale, rounding=ROUND_HALF_UP))


def item_header(value: object) -> tuple[str, str]:
    raw = text(value)
    if not raw:
        return "", ""
    name = raw.splitlines()[0].strip()
    match = re.search(r"单位\s*[：:]\s*([^\s]+)", raw)
    unit = match.group(1).strip() if match else "克"
    if name == "新鲜牛油果":
        unit = "个"
    return name, unit


def item_code(name: str) -> str:
    if name in ITEM_CODES:
        return ITEM_CODES[name]
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:12].upper()
    return f"HIST_LOSS_{digest}"


def parse_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, loss_month: str) -> dict[str, object]:
    label_rows = {
        text(sheet.cell(row, 1).value): row
        for row in range(1, sheet.max_row + 1)
        if text(sheet.cell(row, 1).value)
    }
    total_row = label_rows["合计（总量）"]
    store_rows = []
    for row in range(3, total_row):
        store_name = text(sheet.cell(row, 1).value)
        if not store_name:
            continue
        if store_name not in STORE_ALIASES:
            raise SystemExit(f"{sheet.title}: 未配置门店映射 {store_name}")
        store_rows.append({"row": row, "name": store_name, "id": STORE_ALIASES[store_name]})

    items = []
    for column in range(3, sheet.max_column + 1):
        name, unit = item_header(sheet.cell(2, column).value)
        if not name:
            continue
        factor = Decimal("500") if unit == "克" else Decimal("1")
        pricing_unit = "斤" if unit == "克" else unit
        total_quantity = decimal(sheet.cell(total_row, column).value) or ZERO
        priced_row = label_rows.get("总量")
        priced_quantity = decimal(sheet.cell(priced_row, column).value) if priced_row else None
        if priced_quantity is None:
            priced_quantity = total_quantity / factor
        price_row = label_rows.get("价格")
        price = decimal(sheet.cell(price_row, column).value) if price_row else None
        loss_amount = decimal(sheet.cell(label_rows["总计损耗金额"], column).value) or ZERO
        supplier_amount = decimal(sheet.cell(label_rows["厂商赔付金额"], column).value) or ZERO
        borne_amount = decimal(sheet.cell(label_rows["店铺承担"], column).value) or ZERO
        items.append({
            "column": column,
            "name": name,
            "code": item_code(name),
            "unit": unit,
            "pricing_unit": pricing_unit,
            "factor": factor,
            "total_quantity": total_quantity,
            "priced_quantity": priced_quantity,
            "price": price,
            "loss_amount": loss_amount,
            "supplier_amount": supplier_amount,
            "borne_amount": borne_amount,
        })

    values = []
    for store in store_rows:
        for item in items:
            quantity = decimal(sheet.cell(store["row"], item["column"]).value)
            if quantity is not None:
                values.append({"store": store, "item": item, "quantity": quantity})

    def declared(label: str) -> Decimal:
        return decimal(sheet.cell(label_rows[label], 2).value) or ZERO

    declared_loss = declared("总计损耗金额")
    declared_supplier = declared("厂商赔付金额")
    declared_borne = declared("店铺承担")
    detail_loss = sum((item["loss_amount"] for item in items), ZERO)
    detail_supplier = sum((item["supplier_amount"] for item in items), ZERO)
    detail_borne = sum((item["borne_amount"] for item in items), ZERO)
    calculated_borne = declared_loss - declared_supplier
    declared_borne_difference = declared_borne - calculated_borne
    detail_loss_difference = detail_loss - declared_loss
    variances = {
        "declared_borne_difference": declared_borne_difference,
        "detail_loss_difference": detail_loss_difference,
        "detail_supplier_difference": detail_supplier - declared_supplier,
        "detail_borne_difference": detail_borne - declared_borne,
    }
    status = "SOURCE_VARIANCE" if any(abs(value) >= Decimal("0.005") for value in variances.values()) else "MATCHED"
    note = "; ".join(
        f"{key}={number(value)}" for key, value in variances.items() if abs(value) >= Decimal("0.005")
    ) or "源表声明总额与明细汇总一致"
    return {
        "id": f"DLMA-{loss_month.replace('-', '')}",
        "month": loss_month,
        "sheet": sheet.title,
        "title": text(sheet.cell(1, 1).value),
        "stores": store_rows,
        "items": items,
        "values": values,
        "declared_loss": declared_loss,
        "detail_loss": detail_loss,
        "declared_supplier": declared_supplier,
        "detail_supplier": detail_supplier,
        "declared_borne": declared_borne,
        "detail_borne": detail_borne,
        "calculated_borne": calculated_borne,
        "declared_borne_difference": declared_borne_difference,
        "detail_loss_difference": detail_loss_difference,
        "status": status,
        "note": note,
    }


def build_sql(archives: list[dict[str, object]], tenant_id: int, imported_by: int | None,
              file_name: str, file_sha256: str) -> str:
    sql = ["start transaction;"]
    for store_id, store_name in HISTORICAL_STORES.items():
        sql.append(
            "insert into store_branch(id, tenant_id, brand_id, code, name, area, region_code, "
            "supply_warehouse_id, status, note) values "
            f"({q(store_id)}, {tenant_id}, 1, {q(store_id)}, {q(store_name)}, '荆州历史门店', "
            f"'JINGZHOU', 1, '已闭店', '历史月度报损源表补录') "
            "on duplicate key update name=values(name), note=values(note);"
        )

    latest_new_items: dict[str, dict[str, object]] = {}
    for archive in sorted(archives, key=lambda row: row["month"]):
        for item in archive["items"]:
            if item["code"].startswith("HIST_LOSS_"):
                current = latest_new_items.get(item["code"])
                if current is None or item["price"] is not None:
                    latest_new_items[item["code"]] = item
    for item in latest_new_items.values():
        price = item["price"] if item["price"] is not None else ZERO
        sql.append(
            "insert into loss_item_config(tenant_id, item_code, item_name, category, unit, pricing_unit, "
            "quantity_per_pricing_unit, unit_price, source_sheet, active) values "
            f"({tenant_id}, {q(item['code'])}, {q(item['name'])}, '历史月度报损', {q(item['unit'])}, "
            f"{q(item['pricing_unit'])}, {number(item['factor'], QUANTITY_SCALE)}, {number(price, QUANTITY_SCALE)}, "
            "'历史月度报损', 1) on duplicate key update item_name=values(item_name), unit=values(unit), "
            "pricing_unit=values(pricing_unit), quantity_per_pricing_unit=values(quantity_per_pricing_unit);"
        )

    months = ", ".join(q(archive["month"]) for archive in archives)
    sql.append(f"delete from daily_loss_monthly_archive where tenant_id={tenant_id} and loss_month in ({months});")
    importer = "null" if imported_by is None else str(imported_by)
    for archive in archives:
        sql.append(
            "insert into daily_loss_monthly_archive(" 
            "id, tenant_id, loss_month, source_sheet, source_title, source_file_name, source_file_sha256, "
            "declared_total_loss_amount, detail_total_loss_amount, declared_supplier_compensation_amount, "
            "detail_supplier_compensation_amount, declared_store_borne_amount, detail_store_borne_amount, "
            "calculated_store_borne_amount, declared_borne_difference, detail_loss_difference, store_count, "
            "item_count, reconciliation_status, source_note, imported_by) values "
            f"({q(archive['id'])}, {tenant_id}, {q(archive['month'])}, {q(archive['sheet'])}, "
            f"{q(archive['title'])}, {q(file_name)}, {q(file_sha256)}, {number(archive['declared_loss'])}, "
            f"{number(archive['detail_loss'])}, {number(archive['declared_supplier'])}, "
            f"{number(archive['detail_supplier'])}, {number(archive['declared_borne'])}, "
            f"{number(archive['detail_borne'])}, {number(archive['calculated_borne'])}, "
            f"{number(archive['declared_borne_difference'])}, {number(archive['detail_loss_difference'])}, "
            f"{len(archive['stores'])}, {len(archive['items'])}, {q(archive['status'])}, {q(archive['note'])}, {importer});"
        )
        for store in archive["stores"]:
            sql.append(
                "insert into daily_loss_monthly_archive_store(archive_id, source_row, store_id, store_name_snapshot) values "
                f"({q(archive['id'])}, {store['row']}, {q(store['id'])}, {q(store['name'])});"
            )
        for item in archive["items"]:
            config_id = (
                f"(select id from loss_item_config where tenant_id={tenant_id} and item_code={q(item['code'])})"
            )
            sql.append(
                "insert into daily_loss_monthly_archive_item(" 
                "archive_id, source_column, item_config_id, item_name_snapshot, input_unit_snapshot, "
                "pricing_unit_snapshot, quantity_per_pricing_unit_snapshot, total_loss_quantity, priced_quantity, "
                "unit_price, unit_price_source, total_loss_amount, supplier_compensation_amount, store_borne_amount) values "
                f"({q(archive['id'])}, {item['column']}, {config_id}, {q(item['name'])}, {q(item['unit'])}, "
                f"{q(item['pricing_unit'])}, {number(item['factor'], QUANTITY_SCALE)}, "
                f"{number(item['total_quantity'], QUANTITY_SCALE)}, {number(item['priced_quantity'], QUANTITY_SCALE)}, "
                f"{number(item['price'], QUANTITY_SCALE)}, {q('SOURCE' if item['price'] is not None else 'NOT_PROVIDED')}, "
                f"{number(item['loss_amount'])}, {number(item['supplier_amount'])}, {number(item['borne_amount'])});"
            )
        for value in archive["values"]:
            item = value["item"]
            store = value["store"]
            config_id = (
                f"(select id from loss_item_config where tenant_id={tenant_id} and item_code={q(item['code'])})"
            )
            sql.append(
                "insert into daily_loss_monthly_archive_store_item(" 
                "archive_id, source_row, source_column, store_id, item_config_id, loss_quantity) values "
                f"({q(archive['id'])}, {store['row']}, {item['column']}, {q(store['id'])}, {config_id}, "
                f"{number(value['quantity'], QUANTITY_SCALE)});"
            )
    sql.append("commit;")
    return "\n".join(sql) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate historical monthly loss archive SQL")
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--tenant-id", type=int, default=1)
    parser.add_argument("--imported-by", type=int, default=3)
    parser.add_argument("--out", required=True)
    parser.add_argument("--audit-out", required=True)
    args = parser.parse_args()

    source = Path(args.xlsx)
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    missing = [sheet for sheet in SHEET_MONTHS if sheet not in workbook.sheetnames]
    if missing:
        raise SystemExit(f"缺少工作表: {', '.join(missing)}")
    archives = [parse_sheet(workbook[sheet], month) for sheet, month in SHEET_MONTHS.items()]
    file_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    Path(args.out).write_text(
        build_sql(archives, args.tenant_id, args.imported_by, source.name, file_hash), encoding="utf-8"
    )
    audit = [{
        "month": archive["month"],
        "sheet": archive["sheet"],
        "stores": len(archive["stores"]),
        "items": len(archive["items"]),
        "values": len(archive["values"]),
        "declared_total_loss": number(archive["declared_loss"]),
        "detail_total_loss": number(archive["detail_loss"]),
        "declared_supplier_compensation": number(archive["declared_supplier"]),
        "declared_store_borne": number(archive["declared_borne"]),
        "calculated_store_borne": number(archive["calculated_borne"]),
        "status": archive["status"],
        "note": archive["note"],
    } for archive in archives]
    Path(args.audit_out).write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
