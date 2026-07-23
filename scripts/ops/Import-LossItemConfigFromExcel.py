#!/usr/bin/env python
"""Generate idempotent loss-item pricing SQL from an approved Excel workbook.

Both the legacy per-input-unit price workbook and the real monthly loss matrix are
supported. Dated loss values are intentionally never imported into configuration.
Zero is a valid price and means the store may order the item from the warehouse free.
"""

from __future__ import annotations

import argparse
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


LEGACY_SHEETS = ("每日报损表", "水果检查表")


def sql_string(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def normalized_name(raw: object) -> str:
    text = "" if raw is None else str(raw)
    text = re.sub(r"[\r\n]+", " ", text)
    text = re.sub(r"(?:[（(]\s*)?单位\s*[:：]\s*[^\s）)]+[）)]?", "", text)
    return re.sub(r"\s+", "", text).strip()


def unit_from_label(raw: object, item_name: str) -> str:
    text = "" if raw is None else str(raw)
    match = re.search(r"单位\s*[:：]\s*([^\s）)]+)", text)
    if match:
        return match.group(1).strip()
    return "个" if item_name == "新鲜牛油果" else "克"


def decimal_price(raw: object) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw)).quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError):
        return None


def pricing_standard(unit: str, price: Decimal, legacy_per_input_unit: bool) -> tuple[str, Decimal, Decimal]:
    if unit == "克":
        normalized_price = price * Decimal("500") if legacy_per_input_unit else price
        return "斤", Decimal("500.0000"), normalized_price.quantize(Decimal("0.0001"))
    return unit, Decimal("1.0000"), price.quantize(Decimal("0.0001"))


def row_data(
    tenant_id: int,
    code: str,
    name: str,
    category: str,
    unit: str,
    price: Decimal,
    source_sheet: str,
    legacy_per_input_unit: bool,
) -> dict[str, object]:
    pricing_unit, factor, normalized_price = pricing_standard(unit, price, legacy_per_input_unit)
    return {
        "tenant_id": tenant_id,
        "item_code": code,
        "item_name": name,
        "category": category,
        "unit": unit,
        "pricing_unit": pricing_unit,
        "quantity_per_pricing_unit": factor,
        "unit_price": normalized_price,
        "source_sheet": source_sheet,
    }


def extract_legacy(workbook: openpyxl.Workbook, tenant_id: int) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sheet_name in LEGACY_SHEETS:
        sheet = workbook[sheet_name]
        prefix = "DAILY_LOSS" if sheet_name == "每日报损表" else "FRUIT_CHECK"
        for column in range(1, sheet.max_column + 1):
            raw_name = sheet.cell(2, column).value
            name = normalized_name(raw_name)
            price = decimal_price(sheet.cell(35, column).value)
            if not name or name == "日期" or price is None:
                continue
            unit = unit_from_label(raw_name, name)
            rows.append(row_data(
                tenant_id, f"{prefix}_{column:03d}", name, sheet_name, unit, price, sheet_name, True
            ))
    return rows


def extract_monthly(
    workbook: openpyxl.Workbook, sheet_name: str, tenant_id: int
) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    price_row = next(
        (row for row in range(1, sheet.max_row + 1) if normalized_name(sheet.cell(row, 1).value) == "价格"),
        None,
    )
    if price_row is None:
        raise SystemExit(f"工作表 {sheet_name} 缺少价格汇总行")
    rows: list[dict[str, object]] = []
    for column in range(2, sheet.max_column + 1):
        raw_name = sheet.cell(2, column).value
        name = normalized_name(raw_name)
        price = decimal_price(sheet.cell(price_row, column).value)
        if not name or name == "合计" or price is None:
            continue
        unit = unit_from_label(raw_name, name)
        rows.append(row_data(
            tenant_id, f"MONTHLY_LOSS_{column:03d}", name, "月度报损", unit, price, sheet_name, False
        ))
    return rows


def extract_rows(xlsx_path: Path, tenant_id: int, sheet_name: str | None) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise SystemExit(f"缺少工作表: {sheet_name}")
        return extract_monthly(workbook, sheet_name, tenant_id)
    if all(name in workbook.sheetnames for name in LEGACY_SHEETS):
        return extract_legacy(workbook, tenant_id)
    candidates = [name for name in workbook.sheetnames if "月份" in name or name.endswith("月")]
    if not candidates:
        raise SystemExit("未找到旧版价格表或月度报损工作表，请使用 --sheet 指定")
    return extract_monthly(workbook, candidates[-1], tenant_id)


def build_sql(rows: list[dict[str, object]]) -> str:
    values = []
    for row in rows:
        values.append("(" + ", ".join([
            str(row["tenant_id"]),
            sql_string(str(row["item_code"])),
            sql_string(str(row["item_name"])),
            sql_string(str(row["category"])),
            sql_string(str(row["unit"])),
            sql_string(str(row["pricing_unit"])),
            str(row["quantity_per_pricing_unit"]),
            str(row["unit_price"]),
            sql_string(str(row["source_sheet"])),
            "1", "current_timestamp", "current_timestamp",
        ]) + ")")
    return """insert into loss_item_config(
  tenant_id, item_code, item_name, category, unit, pricing_unit,
  quantity_per_pricing_unit, unit_price, source_sheet, active, created_at, updated_at
)
values
  {values}
on duplicate key update
  item_name = values(item_name), category = values(category), unit = values(unit),
  pricing_unit = values(pricing_unit),
  quantity_per_pricing_unit = values(quantity_per_pricing_unit),
  unit_price = values(unit_price), source_sheet = values(source_sheet),
  active = values(active), updated_at = current_timestamp;
""".format(values=",\n  ".join(values))


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate loss_item_config SQL from Excel pricing rows.")
    parser.add_argument("--xlsx", required=True, help="Approved Excel workbook path")
    parser.add_argument("--tenant-id", type=int, default=1, help="Target tenant id, default 1")
    parser.add_argument("--sheet", help="Monthly sheet, for example 10月份")
    parser.add_argument("--out", help="Output .sql path; prints to stdout when omitted")
    args = parser.parse_args()
    rows = extract_rows(Path(args.xlsx), args.tenant_id, args.sheet)
    sql = build_sql(rows)
    if args.out:
        Path(args.out).write_text(sql, encoding="utf-8")
    else:
        print(sql)
    print(f"-- extracted {len(rows)} loss item config rows", flush=True)


if __name__ == "__main__":
    main()
